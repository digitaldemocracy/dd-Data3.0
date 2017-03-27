#!/usr/bin/env python
'''
File: insert_Behests.py
Author: Mandy Chan
Date: 7/18/2015

Description:
  - Gathers Behest Data and puts it into DDDB

NOTE: If you want to use this and you grab the data from Windows and vim shows 
      ^M instead of line breaks, use the following command:
      
      :%s/^V^M/\r/g (where ^V^M means CTRL+V, CTRL+M)

      in order to replace all instances of ^M with newline characters. 
      Otherwise, this script will think there is only one line in the file.

Usage:
  python insert_Behests.py [file_name.csv]
  - file_name.csv : a .csv file containing the Behests

Source:
  - California Fair Political Practices Commission (fppc.ca.gov/index.php?id=499)
    - YYYY-Senate.xlsx
    - YYYY-Assembly.xlsx

Populates:
  - Behests (official, datePaid, payor, amount, payee, description, purpose, noticeReceived)
  - Organizations (name, city, state)
  - Payors (name, city, state)

'''

import json
import os
import re
import sys
import csv
import requests
import MySQLdb
import openpyxl
import traceback
from pprint import pprint
from datetime import datetime
from Database_Connection import mysql_connection
from graylogger.graylogger import GrayLogger

# Global Data
#GreyLogger info
API_URL = "http://dw.digitaldemocracy.org:12202/gelf"
logger = None
B_INSERT = 0

#URLS for behest data
ASSEMBLY_URL = "http://www.fppc.ca.gov/content/dam/fppc/documents/behested-payments/2016/2016_Assembly.xlsx"

SENATE_URL = "http://www.fppc.ca.gov/content/dam/fppc/documents/behested-payments/2016/2016_Senate.xlsx"

# Number of cmdline arguments needed
NUM_ARGS = 2

# Column names dictionary
COL = {
  'official': 0, 'pay_date': 1, 'payor': 2, 'payor_city': 3, 'payor_state': 4,
  'amount': 5,   'payee': 6,    'payee_city': 7,  'payee_state': 8,
  'descr': 9,    'purpose': 10, 'notice_rec': 11, 'ytd': 12
}

# Dictionary of problematic legislator names. Add as necessary.
NAME_EXCEPTIONS = {
    "Achadjian, Katcho":"Achadjian, K.H. \"Katcho\"",
    "Allen, Ben":"Allen, Benjamin",
    "Bonilla, Susan A.":"Bonilla, Susan",
    "Calderon, Charles":"Calderon, Ian Charles",
    "Calderon, Ian":"Calderon, Ian Charles",
    "Chau, Edwin":"Chau, Ed",
    "DeLeon, Kevin":"De Leon, Kevin",
    "Eggman, Susan":"Eggman, Susan Talamantes",
    "Frazier, James":"Frazier, Jim",
    "Hall, Isadore III":"Hall, Isadore",
    "Jackson, Hanna- Beth":"Jackson, Hannah-Beth",
    "Jones-Sawyer, Reginald Byron":"Jones-Sawyer, Reginald",
    "Perea, Henry T.":"Perea, Henry",
    "Rodriquez, Freddie":"Rodriguez, Freddie",
    "Stone, Jeffrey":"Stone, Jeff",
    "Thomas-Ridley, Sebastian":"Ridley-Thomas, Sebastian",
    "Ting, Phil":"Ting, Philip",
    "Vidak, James Andy":"Vidak, Andy",
}

def create_payload(table, sqlstmt):
  return {'_table':table,
          '_sqlstmt':sqlstmt,
          '_state':'CA'}

'''
Finds the pid of the official. If found, returns tuple (name, pid). Otherwise,
(name, -1).
'''
def find_official(dd_cursor, name):
  # Check and refactor legislator name if necessary
  name = NAME_EXCEPTIONS.get(name, name)

  # Find legislator pid from name
  select_stmt = '''SELECT l.pid
                   FROM Person p
                   JOIN Legislator l ON p.pid = l.pid
                   WHERE CONCAT_WS(', ',last,first) = %(name)s;
                '''
  dd_cursor.execute(select_stmt, {'name':name})

  # Returns the pid if found, otherwise -1
  query = dd_cursor.fetchone()
  if query is None:
    return (name, -1)
  return (name, query[0])

'''
Creates a row in Organizations table.
'''
def create_organization(dd_cursor, org, city, state):
  insert_stmt = '''INSERT INTO Organizations
                   (name, city, stateHeadquartered)
                   VALUES
                   (%(name)s, %(city)s, %(state)s);
                '''
  
  temp = {'name': org, 'city': city, 'state':state}

  try:
    dd_cursor.execute(insert_stmt, temp)
  except:
    logger.warning("Insert statement failed.", full_msg = traceback.format_exc(),
     additional_fields = create_payload('Organizations', (insert_stmt % temp)))
  
  return dd_cursor.lastrowid

'''
Finds organization from the organization name, city, and state. If found, 
returns the organization's 'oid'. Otherwise, creates an organization row and 
returns the newly created 'oid'.
'''
def find_organization(dd_cursor, org, city, state):
  select_stmt = '''SELECT oid
                   FROM Organizations
                   WHERE name = %(name)s
                    AND city = %(city)s
                    AND stateHeadquartered = %(state)s
                '''
  dd_cursor.execute(select_stmt, {'name':org, 'city':city, 'state':state})
  query = dd_cursor.fetchone()
  if query is None:
    return create_organization(dd_cursor, org, city, state)
  return query[0]

'''
Creates a row in Payors table.
'''
def create_payor(dd_cursor, payor, city, state):
  insert_stmt = '''INSERT INTO Payors (name, city, state)
                   VALUES (%(name)s, %(city)s, %(state)s);
                '''
  temp = {'name':payor, 'city':city, 'state':state}
    
  try:
    dd_cursor.execute(insert_stmt, temp)
  except:
    logger.warning("Insert statement failed.", full_msg = traceback.format_exc(),
     additional_fields = create_payload('Payors', (insert_stmt % temp)))
  return dd_cursor.lastrowid

'''
Finds payor from the payor name, city, and state. If found, returns the payor's
'prid'. Otherwise, create a payor row and return the newly created 'prid'.
'''
def find_payor(dd_cursor, payor, city, state):
  select_stmt = '''SELECT prid
                   FROM Payors
                   WHERE name = %(name)s
                    AND city = %(city)s
                    AND state = %(state)s
                '''
  dd_cursor.execute(select_stmt, {'name':payor, 'city':city, 'state':state})
  query = dd_cursor.fetchone()
  if query is None:
    return create_payor(dd_cursor, payor, city, state)
  return query[0]

'''
Check if there is already a behest with the same exact information
'''
def duplicate_behest(dd_cursor,
    off_pid, date_paid, payor_id, amt, payee_id, descr, purpose, notice_rec):
  #select_stmt = '''SELECT *
  #                 FROM Behests
  #                 WHERE official = %(off_pid)s
  #                  AND datePaid = %(date)s AND payor = %(payor)s
  #                  AND amount = %(amt)s AND payee = %(payee)s
  #                  AND description = %(descr)s AND purpose = %(purpose)s
  #                  AND noticeReceived = %(notice_rec)s
  #              '''
  #dd_cursor.execute(select_stmt, {'off_pid':off_pid, 'date':date_paid, 
  #  'payor':payor_id, 'amt':amt, 'payee':payee_id, 'descr':descr,
  #  'purpose':purpose, 'notice_rec':notice_rec})

  select_stmt = '''SELECT *
                   FROM Behests
                   WHERE official = %(off_pid)s
                    AND datePaid = %(date)s AND payor = %(payor)s
                    AND payee = %(payee)s'''

  dd_cursor.execute(select_stmt, {'off_pid':off_pid, 'date':date_paid,
    'payor':payor_id, 'payee':payee_id})

  query = dd_cursor.fetchone()
  if query is None:
    return False
  return True

def get_session_year(dd_cursor):
    select_stmt = '''SELECT MAX(start_year)
                     FROM Session
                     WHERE state = 'CA'
                  '''
    dd_cursor.execute(select_stmt)
    year = dd_cursor.fetchone()[0]
    return year

'''
Insert row into Behests
'''
def create_behest(dd_cursor,
    off_pid, date_paid, payor_id, amt, payee_id, descr, purpose, notice_rec):
  if duplicate_behest(dd_cursor, off_pid, date_paid, payor_id, amt, payee_id,
      descr, purpose, notice_rec):
    return 0

  session_year = get_session_year(dd_cursor)

  insert_stmt = '''INSERT INTO Behests
                   (official, datePaid, payor, amount, payee, description, purpose, noticeReceived, sessionYear)
                   VALUES
                   (%(off_pid)s, %(datePaid)s, %(payor_id)s, %(amount)s, 
                     %(payee_id)s, %(descr)s, %(purpose)s, %(notice_rec)s, %(session_year)s)
                ''' 
  
  temp = {'off_pid':off_pid, 'datePaid':date_paid, 'payor_id':payor_id,
    'amount':amt, 'payee_id':payee_id, 'descr':descr, 'purpose':purpose,
    'notice_rec':notice_rec, 'session_year':session_year}

  try:
    dd_cursor.execute(insert_stmt, temp)
  except:
    logger.warning("Insert statement failed.", full_msg = traceback.format_exc(),
     additional_fields = create_payload('Behests', (insert_stmt % temp)))

  return dd_cursor.rowcount

'''
Parse the row and insert the information to DDDB. It then returns the current 
official because the Behest files don't have the officials mentioned every 
line. (see Behest data)
'''
def parse_row(dd_cursor, attribs, official):
  global B_INSERT

  # Assume if Official and Payor are blank, the line is unnecessary
  if attribs[COL['official']].value == None and attribs[COL['payor']].value == None:
    return official
  # If legislator/official name is present in this row, find their pid
  elif attribs[COL['official']].value != None:
    official = find_official(dd_cursor, attribs[COL['official']].value.strip())
  
  name = official[0]
  pid = official[1]

  # If official isn't found, return without inserting anything
  if pid < 0:
    print('Could not find Legislator: %(name)s, skipping' % {'name':name})
    return (name, pid)

  # Find Payor id from Payors table
  payor_id = find_payor(dd_cursor,
      attribs[COL['payor']].value.strip(), attribs[COL['payor_city']].value.strip(),
      attribs[COL['payor_state']].value.strip())

  # Find Payee id from Organizations table 
  payee_id = find_organization(dd_cursor,
      attribs[COL['payee']].value.strip(), attribs[COL['payee_city']].value.strip(),
      attribs[COL['payee_state']].value.strip())

  # Rest of variables
  date_paid = attribs[COL['pay_date']].value
  amount = attribs[COL['amount']].value
  descr = attribs[COL['descr']].value
  purpose = attribs[COL['purpose']].value
  notice_rec = None if attribs[COL['notice_rec']].value == 'No date' or type(attribs[COL['notice_rec']].value) is not datetime else (
      attribs[COL['notice_rec']].value)

  # Format dates correctly
  try:
      date_paid = date_paid.strftime('%Y-%m-%d')
  except AttributeError:
    #pattern = '\d{1,2}/\d{1,2}-\d{1,2}/\d{1,2}/\d{2}'
    #if re.match(pattern, date_paid):
    date_paid = date_paid.split('-')[0].strip()
    try:
      date_paid = (datetime.strptime(date_paid, '%m/%d/%y').
       strftime('%Y-%m-%d'))
    except ValueError:
      date_paid = (datetime.strptime(date_paid, '%m/%Y').
       strftime('%Y-%m-%d'))
    #else:
    #  print('Incorrect date format! %s' % date_paid)
 
  B_INSERT += create_behest(dd_cursor, pid, date_paid, payor_id, amount, payee_id, descr, 
      purpose, notice_rec)
  return official

'''
Downloads behest information from the FPPC website
'''
def download_behests():
  with open("senate_behests.xlsx", "wb") as file:
    response = requests.get(SENATE_URL)
    file.write(response.content)
    file.close

  with open("assembly_behests.xlsx", "wb") as file:
    response = requests.get(ASSEMBLY_URL)
    file.write(response.content)
    file.close

def main():
  global logger
  # Current Official (as (name, pid) tuple)
  cur_official = (None, -1)

  download_behests()

  dbinfo = mysql_connection(sys.argv)

  # Opening db connection
  with MySQLdb.connect(host = dbinfo['host'],
                       port = dbinfo['port'],
                       db = dbinfo['db'],
                       user = dbinfo['user'],
                       passwd = dbinfo['passwd']) as dd_cursor:
    with GrayLogger(API_URL) as _logger:
      logger = _logger  
      
      # Opening file and start at the header_line number
      assembly = openpyxl.load_workbook("assembly_behests.xlsx", data_only = True)
      assembly_ws = assembly['Sheet1']

      for row in assembly_ws.iter_rows(min_row = 6, max_row = assembly_ws.max_row):
        cur_official = parse_row(dd_cursor, row, cur_official)

      senate = openpyxl.load_workbook("senate_behests.xlsx", data_only = True)
      senate_ws = senate['Sheet1']

      for row in senate_ws.iter_rows(min_row = 6, max_row = senate_ws.max_row):
        cur_official = parse_row(dd_cursor, row, cur_official)

      logger.info(__file__ + " terminated successfully.",
        full_msg = "Inserted " + str(B_INSERT) + " rows in Behests.",
        additional_fields = {'_affected_rows': "Behests: " + str(B_INSERT),
                             '_inserted': "Behests: " + str(B_INSERT),
                             '_state': 'CA',
                             '_log_type': 'Database'})

  LOG = {'tables': [{'state': 'CA', 'name': 'Behests', 'inserted':B_INSERT, 'updated': 0, 'deleted': 0}]}
  sys.stderr.write(json.dumps(LOG))

if __name__ == "__main__":
  main()
